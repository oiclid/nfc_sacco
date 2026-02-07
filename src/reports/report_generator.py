"""
Reports Engine - Generate PDF and Excel reports
===============================================
"""

from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os


class ReportGenerator:
    """Generate various reports in PDF and Excel formats"""
    
    def __init__(self, db_manager):
        self.db = db_manager
        self.reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'data', 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    # ========================================================================
    # PDF REPORTS
    # ========================================================================
    
    def generate_member_statement_pdf(self, member_id, start_date, end_date):
        """Generate member statement PDF"""
        # Get member info
        member = self.db.get_member(member_id)
        if not member:
            raise ValueError(f"Member {member_id} not found")
        
        # Create filename
        filename = f"Member_Statement_{member_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        filepath = os.path.join(self.reports_dir, filename)
        
        # Create PDF
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2980B9'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Organization header
        org_name = self.db.get_setting('organization_name') or 'NFC Cooperative'
        story.append(Paragraph(org_name, title_style))
        story.append(Paragraph("MEMBER ACCOUNT STATEMENT", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        # Member info
        full_name = f"{member['first_name']} "
        if member['middle_name']:
            full_name += f"{member['middle_name']} "
        full_name += member['last_name']
        
        member_info = [
            ['Member ID:', member_id],
            ['Name:', full_name],
            ['Date:', datetime.now().strftime('%B %d, %Y')],
            ['Period:', f"{start_date} to {end_date}"]
        ]
        
        member_table = Table(member_info, colWidths=[2*inch, 4*inch])
        member_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(member_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Account Summary
        summary = self.db.get_member_summary(member_id)
        if summary:
            summary = summary[0]
            
            story.append(Paragraph("ACCOUNT SUMMARY", styles['Heading3']))
            
            summary_data = [
                ['Description', 'Amount (₦)'],
                ['Total Savings', f"{summary['total_savings']:,.2f}"],
                ['  - Premium Savings', f"{summary['premium_savings']:,.2f}"],
                ['  - Fixed/Target Deposits', f"{summary['fixed_target_deposits']:,.2f}"],
                ['  - Share Investment', f"{summary['shares_investment']:,.2f}"],
                ['Total Loans Outstanding', f"{summary['total_loans_outstanding']:,.2f}"],
                ['Net Balance', f"{summary['net_balance']:,.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2980B9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Transactions
        story.append(Paragraph("TRANSACTION HISTORY", styles['Heading3']))
        
        transactions = self.db.get_transactions(member_id, start_date, end_date)
        
        if transactions:
            trans_data = [['Date', 'Type', 'Description', 'Debit (₦)', 'Credit (₦)']]
            
            for txn in transactions:
                debit = f"{txn['amount']:,.2f}" if not txn['is_credit'] else '-'
                credit = f"{txn['amount']:,.2f}" if txn['is_credit'] else '-'
                
                trans_data.append([
                    txn['transaction_date'],
                    txn['transaction_type'],
                    txn['description'] or '',
                    debit,
                    credit
                ])
            
            trans_table = Table(trans_data, colWidths=[1*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
            trans_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            story.append(trans_table)
        else:
            story.append(Paragraph("No transactions found for this period.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        return filepath
    
    def generate_cashbook_pdf(self, start_date, end_date):
        """Generate cashbook report"""
        filename = f"Cashbook_{start_date}_to_{end_date}.pdf"
        filepath = os.path.join(self.reports_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=18, textColor=colors.HexColor('#2980B9'),
                                     alignment=TA_CENTER)
        
        org_name = self.db.get_setting('organization_name') or 'NFC Cooperative'
        story.append(Paragraph(org_name, title_style))
        story.append(Paragraph("CASHBOOK REPORT", styles['Heading2']))
        story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Get all transactions
        transactions = self.db.get_transactions(None, start_date, end_date)
        
        # Group by payment method
        cash_transactions = [t for t in transactions if t.get('payment_method') == 'Cash']
        
        if cash_transactions:
            data = [['Date', 'Member ID', 'Description', 'Receipts (₦)', 'Payments (₦)']]
            
            total_receipts = 0
            total_payments = 0
            
            for txn in cash_transactions:
                if txn['is_credit']:
                    receipts = f"{txn['amount']:,.2f}"
                    payments = '-'
                    total_receipts += txn['amount']
                else:
                    receipts = '-'
                    payments = f"{txn['amount']:,.2f}"
                    total_payments += txn['amount']
                
                data.append([
                    txn['transaction_date'],
                    txn['member_id'],
                    txn['description'] or txn['transaction_type'],
                    receipts,
                    payments
                ])
            
            # Totals
            data.append(['', '', 'TOTALS:', f"{total_receipts:,.2f}", f"{total_payments:,.2f}"])
            
            table = Table(data, colWidths=[1*inch, 1*inch, 2.5*inch, 1.2*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2980B9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No cash transactions found for this period.", styles['Normal']))
        
        doc.build(story)
        return filepath
    
    # ========================================================================
    # EXCEL REPORTS
    # ========================================================================
    
    def generate_member_summary_excel(self, start_date, end_date):
        """Generate member summary in Excel"""
        filename = f"Member_Summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Member Summary"
        
        # Styles
        header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        currency_format = '₦#,##0.00'
        
        # Title
        ws['A1'] = self.db.get_setting('organization_name') or 'NFC Cooperative'
        ws['A1'].font = title_font
        ws['A2'] = 'MEMBER SUMMARY REPORT'
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f'As at: {datetime.now().strftime("%B %d, %Y")}'
        
        # Headers
        headers = ['Member ID', 'Name', 'Premium Savings', 'Fixed/Target Deposits',
                  'Share Investment', 'Total Savings', 'Loans Outstanding', 'Net Balance']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get member summaries
        summaries = self.db.get_member_summary()
        
        # Data rows
        row = 6
        for summary in summaries:
            ws.cell(row=row, column=1, value=summary['member_id'])
            ws.cell(row=row, column=2, value=summary['full_name'])
            ws.cell(row=row, column=3, value=summary['premium_savings']).number_format = currency_format
            ws.cell(row=row, column=4, value=summary['fixed_target_deposits']).number_format = currency_format
            ws.cell(row=row, column=5, value=summary['shares_investment']).number_format = currency_format
            ws.cell(row=row, column=6, value=summary['total_savings']).number_format = currency_format
            ws.cell(row=row, column=7, value=summary['total_loans_outstanding']).number_format = currency_format
            ws.cell(row=row, column=8, value=summary['net_balance']).number_format = currency_format
            row += 1
        
        # Totals row
        total_row = row
        ws.cell(row=total_row, column=1, value='TOTALS').font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f'{len(summaries)} Members').font = Font(bold=True)
        
        for col in range(3, 9):
            formula = f'=SUM({ws.cell(row=6, column=col).coordinate}:{ws.cell(row=total_row-1, column=col).coordinate})'
            cell = ws.cell(row=total_row, column=col, value=formula)
            cell.number_format = currency_format
            cell.font = Font(bold=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18
        
        # Save
        wb.save(filepath)
        return filepath
    
    def generate_loan_portfolio_excel(self):
        """Generate loan portfolio analysis in Excel"""
        filename = f"Loan_Portfolio_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Portfolio Summary"
        
        # Get all loans
        all_loans = self.db.fetchall("SELECT * FROM loans")
        active_loans = [l for l in all_loans if l['status'] == 'Active']
        
        # Headers
        header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws_summary['A1'] = 'LOAN PORTFOLIO ANALYSIS'
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A2'] = f'As at: {datetime.now().strftime("%B %d, %Y")}'
        
        # Summary statistics
        summary_data = [
            ['Metric', 'Value'],
            ['Total Loans Disbursed', len(all_loans)],
            ['Active Loans', len(active_loans)],
            ['Completed Loans', len([l for l in all_loans if l['status'] == 'Completed'])],
            ['Total Amount Disbursed', sum(l['principal_amount'] for l in all_loans)],
            ['Total Outstanding', sum(l['balance_outstanding'] for l in active_loans)],
            ['Total Collected', sum(l['amount_paid'] for l in all_loans)],
        ]
        
        for row, (label, value) in enumerate(summary_data, 4):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws_summary.cell(row=row, column=2, value=value)
            if isinstance(value, (int, float)) and row > 4:
                cell.number_format = '₦#,##0.00'
        
        # Detailed loans sheet
        ws_details = wb.create_sheet("Loan Details")
        
        headers = ['Loan Number', 'Member ID', 'Type', 'Principal', 'Interest',
                  'Total Amount', 'Amount Paid', 'Balance', 'Status', 'Start Date']
        
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, loan in enumerate(all_loans, 2):
            ws_details.cell(row=row, column=1, value=loan['loan_number'])
            ws_details.cell(row=row, column=2, value=loan['member_id'])
            
            # Get loan type name
            loan_type = self.db.fetchone(
                "SELECT type_name FROM loan_types WHERE loan_type_id = ?",
                (loan['loan_type_id'],)
            )
            ws_details.cell(row=row, column=3, value=loan_type['type_name'] if loan_type else '')
            
            ws_details.cell(row=row, column=4, value=loan['principal_amount']).number_format = '₦#,##0.00'
            ws_details.cell(row=row, column=5, value=loan['interest_amount']).number_format = '₦#,##0.00'
            ws_details.cell(row=row, column=6, value=loan['total_amount']).number_format = '₦#,##0.00'
            ws_details.cell(row=row, column=7, value=loan['amount_paid']).number_format = '₦#,##0.00'
            ws_details.cell(row=row, column=8, value=loan['balance_outstanding']).number_format = '₦#,##0.00'
            ws_details.cell(row=row, column=9, value=loan['status'])
            ws_details.cell(row=row, column=10, value=loan['start_date'])
        
        # Auto-size columns
        for ws in [ws_summary, ws_details]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath
